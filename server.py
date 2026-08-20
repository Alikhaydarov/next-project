from __future__ import annotations

import json
import mimetypes
import os
import sqlite3
import threading
from datetime import datetime
from zoneinfo import ZoneInfo
from http import HTTPStatus
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
STATIC_DIR = ROOT / "static"
EXPORT_DIR = ROOT / "exports"
DB_PATH = DATA_DIR / "payments.db"
XLSX_PATH = EXPORT_DIR / "payments.xlsx"
LOCK = threading.Lock()

DATA_DIR.mkdir(exist_ok=True)
EXPORT_DIR.mkdir(exist_ok=True)


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS our_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                label TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS company_accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                label TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(company_id, label),
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                our_account_id INTEGER NOT NULL,
                company_account_id INTEGER NOT NULL,
                paid_at TEXT NOT NULL,
                amount INTEGER NOT NULL CHECK(amount > 0),
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(company_id) REFERENCES companies(id),
                FOREIGN KEY(our_account_id) REFERENCES our_accounts(id),
                FOREIGN KEY(company_account_id) REFERENCES company_accounts(id)
            );
            """
        )


def payment_rows(conn):
    return conn.execute(
        """
        SELECT p.id,
               c.name AS company_name,
               oa.label AS our_account,
               ca.label AS company_account,
               p.paid_at,
               p.amount,
               p.created_at
        FROM payments p
        JOIN companies c ON c.id = p.company_id
        JOIN our_accounts oa ON oa.id = p.our_account_id
        JOIN company_accounts ca ON ca.id = p.company_account_id
        ORDER BY datetime(p.paid_at) DESC, p.id DESC
        """
    ).fetchall()


def regenerate_excel():
    with LOCK:
        with db() as conn:
            rows = payment_rows(conn)

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"

        headers = [
            "Company name",
            "Our Card Account",
            "Company Card Account",
            "Date & Time",
            "Amount",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in rows[::-1]:
            try:
                paid_dt = datetime.fromisoformat(r["paid_at"])
            except ValueError:
                paid_dt = r["paid_at"]
            ws.append([
                r["company_name"],
                r["our_account"],
                r["company_account"],
                paid_dt,
                r["amount"],
            ])

        ws.freeze_panes = "A2"
        widths = [28, 28, 30, 22, 18]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        for cell in ws["D"][1:]:
            cell.number_format = "yyyy-mm-dd hh:mm"
        for cell in ws["E"][1:]:
            cell.number_format = '₩#,##0'

        # Summary sheet
        summary = wb.create_sheet("Summary")
        summary["A1"] = "PAYMENT SUMMARY"
        summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        summary["A1"].fill = PatternFill("solid", fgColor="0F172A")
        summary.merge_cells("A1:B1")
        summary["A3"] = "Total Payments"
        summary["B3"] = f"=SUM(Payments!E2:E{max(2, len(rows)+1)})"
        summary["B3"].number_format = '₩#,##0'
        summary["A4"] = "Payment Records"
        summary["B4"] = len(rows)
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 20

        wb.save(XLSX_PATH)


def get_data():
    with db() as conn:
        companies = [dict(r) for r in conn.execute("SELECT id, name FROM companies ORDER BY name")]
        our_accounts = [dict(r) for r in conn.execute("SELECT id, label FROM our_accounts ORDER BY label")]
        company_accounts = [dict(r) for r in conn.execute(
            "SELECT id, company_id, label FROM company_accounts ORDER BY company_id, label"
        )]
        payments = [dict(r) for r in payment_rows(conn)]

        total = conn.execute("SELECT COALESCE(SUM(amount), 0) AS total FROM payments").fetchone()["total"]
        now_seoul = datetime.now(ZoneInfo("Asia/Seoul"))
        today_prefix = now_seoul.strftime("%Y-%m-%d")
        month_prefix = now_seoul.strftime("%Y-%m")
        today = conn.execute(
            "SELECT COALESCE(SUM(amount),0) AS total FROM payments WHERE substr(paid_at,1,10)=?",
            (today_prefix,),
        ).fetchone()["total"]
        month = conn.execute(
            "SELECT COALESCE(SUM(amount),0) AS total FROM payments WHERE substr(paid_at,1,7)=?",
            (month_prefix,),
        ).fetchone()["total"]

    return {
        "companies": companies,
        "our_accounts": our_accounts,
        "company_accounts": company_accounts,
        "payments": payments,
        "stats": {"total": total, "today": today, "month": month, "count": len(payments)},
    }


class Handler(BaseHTTPRequestHandler):
    server_version = "PaymentAdmin/1.0"

    def log_message(self, fmt, *args):
        print(f"[{self.log_date_time_string()}] {fmt % args}")

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0") or 0)
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def serve_file(self, path: Path, content_type=None, download_name=None):
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type or mimetypes.guess_type(str(path))[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        if download_name:
            self.send_header("Content-Disposition", f'attachment; filename="{download_name}"')
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        path = urlparse(self.path).path
        if path == "/":
            return self.serve_file(STATIC_DIR / "index.html", "text/html; charset=utf-8")
        if path == "/api/data":
            return self.send_json(get_data())
        if path == "/export.xlsx":
            regenerate_excel()
            return self.serve_file(
                XLSX_PATH,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "payments.xlsx",
            )
        if path.startswith("/static/"):
            rel = path[len("/static/"):]
            candidate = (STATIC_DIR / rel).resolve()
            if STATIC_DIR.resolve() not in candidate.parents:
                return self.send_error(403)
            return self.serve_file(candidate)
        self.send_error(404)

    def do_POST(self):
        path = urlparse(self.path).path
        try:
            body = self.read_json()

            if path == "/api/companies":
                name = str(body.get("name", "")).strip()
                if not name:
                    return self.send_json({"error": "Company name is required"}, 400)
                with db() as conn:
                    cur = conn.execute("INSERT INTO companies(name) VALUES (?)", (name,))
                    cid = cur.lastrowid
                return self.send_json({"ok": True, "id": cid}, 201)

            if path == "/api/our-accounts":
                label = str(body.get("label", "")).strip()
                if not label:
                    return self.send_json({"error": "Card account is required"}, 400)
                with db() as conn:
                    cur = conn.execute("INSERT INTO our_accounts(label) VALUES (?)", (label,))
                    aid = cur.lastrowid
                return self.send_json({"ok": True, "id": aid}, 201)

            if path == "/api/company-accounts":
                label = str(body.get("label", "")).strip()
                company_id = int(body.get("company_id") or 0)
                if not label or not company_id:
                    return self.send_json({"error": "Company and card account are required"}, 400)
                with db() as conn:
                    cur = conn.execute(
                        "INSERT INTO company_accounts(company_id,label) VALUES (?,?)",
                        (company_id, label),
                    )
                    aid = cur.lastrowid
                return self.send_json({"ok": True, "id": aid}, 201)

            if path == "/api/payments":
                company_id = int(body.get("company_id") or 0)
                our_account_id = int(body.get("our_account_id") or 0)
                company_account_id = int(body.get("company_account_id") or 0)
                paid_at = str(body.get("paid_at", "")).strip()
                raw_amount = str(body.get("amount", "")).replace(",", "").replace(" ", "").replace("₩", "")
                amount = int(raw_amount or 0)
                if not all([company_id, our_account_id, company_account_id, paid_at, amount > 0]):
                    return self.send_json({"error": "Please complete all fields"}, 400)

                # Make sure selected company account belongs to selected company.
                with db() as conn:
                    valid = conn.execute(
                        "SELECT 1 FROM company_accounts WHERE id=? AND company_id=?",
                        (company_account_id, company_id),
                    ).fetchone()
                    if not valid:
                        return self.send_json({"error": "The selected company account does not belong to this company"}, 400)
                    cur = conn.execute(
                        """
                        INSERT INTO payments(company_id,our_account_id,company_account_id,paid_at,amount)
                        VALUES (?,?,?,?,?)
                        """,
                        (company_id, our_account_id, company_account_id, paid_at, amount),
                    )
                    pid = cur.lastrowid

                regenerate_excel()
                return self.send_json({"ok": True, "id": pid, "download": "/export.xlsx"}, 201)

            return self.send_json({"error": "Not found"}, 404)

        except sqlite3.IntegrityError as exc:
            return self.send_json({"error": "This record already exists or contains invalid data", "detail": str(exc)}, 409)
        except (ValueError, TypeError, json.JSONDecodeError):
            return self.send_json({"error": "Invalid input format"}, 400)
        except Exception as exc:
            print("ERROR", repr(exc))
            return self.send_json({"error": "Server error"}, 500)


def main():
    init_db()
    regenerate_excel()
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"Payment Admin running on http://localhost:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass


if __name__ == "__main__":
    main()
